import datetime
import json
import requests
import openpyxl


class Household:
    def __init__(self):
        self.today = datetime.datetime.now()
        self.file_date_str = '_' + self.today.strftime('%Y%m%d')

        self.domain = 'http://newcloud.itianluo.cn/householdV2/zoneUserInfo.do?{param}'
        self.base_url = self.domain.format(
            param='app=GJ&client=iPhone&version=1.6.4&outerType=STAFF&token=4948999e91a6947e0b0b8c925badaa16&userId=431&zoneId={zoneId}'
        )

        self.zoneName = {
            76: '天水家园',
            82: '浙地人家',
            84: '南城景苑',
            85: '三水一生'
        }

    def handle(self):
        data = list()

        zone_ids = [76, 82, 84, 85]
        for zone_id in zone_ids:
            response = requests.get(self.base_url.format(zoneId=zone_id))
            json_data = json.loads(response.text)

            for allow in json_data['data']['allow']:
                for row in allow['data']:
                    row['zoneName'] = self.zoneName[zone_id]
                    data.append(row)
                    print(row)

        title = '浙地_南城_三水_天水_认证住户'
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title

        sheet.cell(row=1, column=1, value='小区')
        sheet.cell(row=1, column=2, value='房号')
        sheet.cell(row=1, column=3, value='认证用户数')

        for row in range(2, len(data) + 2):
            sheet.cell(row=row, column=1, value=data[row - 2]['zoneName'])
            sheet.cell(row=row, column=2, value=data[row - 2]['fullName'])
            sheet.cell(row=row, column=3, value=data[row - 2]['num'])

        wb.save(title + self.file_date_str + '.xlsx')


if __name__ == '__main__':
    tjMd = Household()
    tjMd.handle()
