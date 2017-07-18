"""
处理菜单
"""
import os

from openpyxl import Workbook

from handle.itl.handleFor170716.menu.data.roleData import RoleData
from utils.file.excel.readUtil import ReadUtil


class ToLegal:
    _TYPE = {'确认': '1A', '拍照': '1B', '检查': '1C', '抽检': '1D', '巡更': '1E'}
    _RATE = {'日': 'D', '周': 'W', '月': 'M', '年': 'Y'}

    def __init__(self, *args, **kw):
        print('init')

    def handle(self, file_name):
        field_index = {
            'role': 0,
            'type': 1,
            'content': 2,
            'standard': 3,
            'rate': 4,
            'startTime': 5,
            'endTime': 6,
        }

        self.write_file(ReadUtil.read_file(file_name, field_index), file_name)

    @staticmethod
    def write_file(data, file_name):
        wb = Workbook()  # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet

        index = 1
        for row in data:
            if ToLegal._check_param(row):
                ws.cell(row=index, column=1).value = ToLegal._type_name_to_type(row['type'])
                ws.cell(row=index, column=2).value = ToLegal._transform_text(row['content'])
                ws.cell(row=index, column=3).value = ToLegal._transform_text(row['standard'])
                ws.cell(row=index, column=4).value = ToLegal._rate_name_to_rate(row['rate'])
                ws.cell(row=index, column=5).value = int(row['startTime'])
                ws.cell(row=index, column=6).value = int(row['endTime'])
                ws.cell(row=index, column=7).value = RoleData.ROLE_ALL[row['role']]
                ws.cell(row=index, column=8).value = row['role']
            else:
                print(row)
            index += 1

        wb.save(filename="/Users/luoyanjie/out_" + file_name)  # 保存

    @staticmethod
    def _check_param(row):
        try:
            start_time = int(row['startTime'])
            end_time = int(row['endTime']) if int(row['endTime']) != -1 else 1

            flag = row['type'] != '' \
                   and row['type'] in ToLegal._TYPE.keys() \
                   and len(row['content'].strip()) <= 256 \
                   and len(row['content'].strip()) <= 512 \
                   and row['rate'] in ToLegal._RATE.keys() \
                   and start_time <= end_time \
                   and row['role'] in RoleData.ROLE_ALL.keys()
            return flag
        except Exception as e:
            print(e)
            return False
        finally:
            pass

    @staticmethod
    def _type_name_to_type(type_name):
        return ToLegal._TYPE[type_name] if type_name in ToLegal._TYPE else ''

    @staticmethod
    def _rate_name_to_rate(rate_name):
        return ToLegal._RATE[rate_name] if rate_name in ToLegal._RATE else ''

    @staticmethod
    def _transform_text(text):
        return str(text).strip().replace(',', '，').replace('\n', '')


if __name__ == '__main__':
    file_name = '房总物业任务池(住宅)-产品导入版.xlsx'

    print(os.getcwd())
    print(os.listdir())

    handle = ToLegal()
    handle.handle(file_name)
