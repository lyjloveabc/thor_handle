import json
import requests
import openpyxl


class Tj:
    def __init__(self):
        self.today_str = '_20180307'

        self.appstore_ios_yz = 'x1nWvW11YltWh17QmU2YR5vQ-gzGzoHsz'
        self.android_yz = 'JM9hpUnSM2c4jn3VHrL5NJGI-gzGzoHsz'

        self.ios = 'ios'
        self.android = 'android'

        self.base_url = 'https://leancloud.cn'
        self.url = '/1/stats/load_table_data?appid={appid}&end_date={ed}&limit={limit}&os={os}&skip=0&start_date={sd}&stats=event_summary'
        self.cookies = {
            'XSRF-TOKEN': 'c8c2662bb42317e376f7fc70e6e5e8198166eaf6f71961b573c42dd00c9b4abc',
            'gr_session_id_a268202b003f2516': '1cd558d3-e4d3-4443-bd67-664d363cde2d',
            'gr_user_id': '1b56188c-6bbe-490d-9c40-9e4da31ae684',
            'uluru_user': 'loXtak8tEi4EnYq58mxYYA%3D%3D',
            '_ga': 'GA1.2.1636345834.1513403984',
            '_gid': 'GA1.2.574324827.1520412993'
        }
        self.limit = 150

    def handle(self):
        app_id = self.android_yz
        sd = '20160801'
        ed = '20180306'
        os = self.android
        title = os + '埋点数据统计'

        response = requests.get(self.base_url + self.url.format(appid=app_id, sd=sd, ed=ed, limit=self.limit, os=os), cookies=self.cookies)

        json_data_init = json.loads(response.text)
        limit = int(json_data_init['total'])
        print('total: ' + str(limit))

        response = requests.get(self.base_url + self.url.format(appid=app_id, sd=sd, ed=ed, limit=limit, os=os), cookies=self.cookies)

        json_data = json.loads(response.text)

        with open(title + self.today_str + '.txt', 'a') as f:
            f.write('事件ID: 事件总数    最后一日事件数')
            f.write('\n')
            for row in json_data['stats']:
                f.write(str(row['event']) + ': ' + str(row['total-pv']) + '    ' + str(row['today-pv']))
                f.write('\n')

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title
        sheet.cell(row=1, column=1, value='事件ID')
        sheet.cell(row=1, column=2, value='事件总数')
        sheet.cell(row=1, column=3, value='最后一日事件数')
        for i in range(2, limit + 2):
            print(str(json_data['stats'][i - 2]['event']))
            sheet.cell(row=i, column=1, value=str(json_data['stats'][i - 2]['event']))
            sheet.cell(row=i, column=2, value=str(json_data['stats'][i - 2]['total-pv']))
            sheet.cell(row=i, column=3, value=str(json_data['stats'][i - 2]['today-pv']))

        wb.save(title + self.today_str + '.xlsx')


if __name__ == '__main__':
    tj = Tj()
    tj.handle()
