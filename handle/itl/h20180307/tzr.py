import datetime
import json
import requests
import openpyxl


class Tzr:
    def __init__(self):
        self.today = datetime.datetime.now()
        self.today_int = int(self.today.strftime('%Y%m%d'))
        self.file_date_str = '_' + str(self.today_int)
        self.url = 'http://support.itianluo.cn/owner/ownerIndexItemTj?token=092ef9338413437abdf58e9560a0842a&day={day}'

    def handle(self):
        init_start_date = '20180101'
        title = '业主服务次数按天统计'
        total_map = dict()
        day_str_list = list()

        start_date = datetime.datetime.strptime(init_start_date, '%Y%m%d')

        while int(start_date.strftime('%Y%m%d')) <= self.today_int:
            day_str = start_date.strftime('%Y-%m-%d')
            day_str_list.append(day_str)

            total_map[day_str] = self.get_data(int(start_date.strftime('%Y%m%d')))
            start_date = start_date + datetime.timedelta(days=1)

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title

        sheet.cell(row=1, column=1, value='日期')
        sheet.cell(row=1, column=2, value='业主报修的总数量')
        sheet.cell(row=1, column=3, value='业主报修已处理的数量')
        sheet.cell(row=1, column=4, value='业主投诉的总数量')
        sheet.cell(row=1, column=5, value='业主投诉已处理的数量')
        sheet.cell(row=1, column=6, value='快递服务总数量')
        sheet.cell(row=1, column=7, value='快递服务入库数量')
        sheet.cell(row=1, column=8, value='快递服务签收数量')
        sheet.cell(row=1, column=9, value='业主评价总数量')
        sheet.cell(row=1, column=10, value='业主评价好评数量')
        sheet.cell(row=1, column=11, value='业主评价差评数量')
        sheet.cell(row=1, column=12, value='表扬总数量')
        sheet.cell(row=1, column=13, value='咨询总数量')
        sheet.cell(row=1, column=14, value='访客物品出入总数量')
        sheet.cell(row=1, column=15, value='总数量')

        row_index = 2
        for key in range(0, len(day_str_list)):
            item = total_map[day_str_list[key]]
            sheet.cell(row=row_index, column=1, value=day_str_list[key])
            sheet.cell(row=row_index, column=2, value=item['repairCount'])
            sheet.cell(row=row_index, column=3, value=item['repairDoneCount'])
            sheet.cell(row=row_index, column=4, value=item['complaintCount'])
            sheet.cell(row=row_index, column=5, value=item['complaintDoneCount'])
            sheet.cell(row=row_index, column=6, value=item['expressCount'])
            sheet.cell(row=row_index, column=7, value=item['expressInCount'])
            sheet.cell(row=row_index, column=8, value=item['expressCheckedCount'])
            sheet.cell(row=row_index, column=9, value=item['evaluationCount'])
            sheet.cell(row=row_index, column=10, value=item['evaluationHighCount'])
            sheet.cell(row=row_index, column=11, value=item['evaluationLowCount'])
            sheet.cell(row=row_index, column=12, value=item['praiseCount'])
            sheet.cell(row=row_index, column=13, value=item['consultationCount'])
            sheet.cell(row=row_index, column=14, value=item['accessCount'])
            sheet.cell(row=row_index, column=15, value=item['total'])
            row_index += 1

        wb.save(title + self.file_date_str + '.xlsx')

    def get_data(self, start_date):
        return json.loads(requests.get(self.url.format(day=start_date)).text)['data']


if __name__ == '__main__':
    tjMd = Tzr()
    tjMd.handle()
