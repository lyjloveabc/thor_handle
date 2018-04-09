import datetime
import json
import requests
import openpyxl


class TjMd:
    def __init__(self):
        self.today = datetime.datetime.now()
        self.file_date_str = '_' + self.today.strftime('%Y%m%d')

        self.appstore_ios_yz = 'x1nWvW11YltWh17QmU2YR5vQ-gzGzoHsz'
        self.android_yz = 'JM9hpUnSM2c4jn3VHrL5NJGI-gzGzoHsz'

        self.ios = 'ios'
        self.android = 'android'

        self.base_url = 'https://leancloud.cn'
        self.url = '/1/stats/load_table_data?appid={appid}&end_date={ed}&limit={limit}&os={os}&skip=0&start_date={sd}&stats=event_summary'
        self.cookies = {
            'XSRF-TOKEN': '1428b2677cf1bcbfdb2e5a2309573eb1d13812d255b8593e68325195cd52013c',
            '_ga': 'GA1.2.1636345834.1513403984',
            '_gat': '1',
            '_gid': 'GA1.2.1820349074.1522845238',
            'gr_session_id_a268202b003f2516': '614820a6-305a-4e68-bcc3-6d001611b293',
            'gr_user_id': '1b56188c-6bbe-490d-9c40-9e4da31ae684',
            'uluru_user': 'wR3SaGf%2Be%2FAOs9ubX2F%2F4g%3D%3D',


        }
        self.limit = 150

    def handle(self):
        app_id = self.appstore_ios_yz
        init_start_date = '20180205'
        os = self.ios
        title = os + '埋点数据按照周统计'
        event_list = list()
        total_map = dict()

        start_date = datetime.datetime.strptime(init_start_date, '%Y%m%d')
        end_date = start_date + datetime.timedelta(days=6)

        for index in range(1, 10):
            total_map[index] = dict()
            items = self.get_data(app_id, start_date.strftime('%Y%m%d'), end_date.strftime('%Y%m%d'), os)

            for item_index in range(0, len(items)):
                if items[item_index]['event'] not in event_list:
                    event_list.append(items[item_index]['event'])
                total_map[index][items[item_index]['event']] = items[item_index]['total-pv']

            start_date = end_date + datetime.timedelta(days=1)
            end_date = start_date + datetime.timedelta(days=6)

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = title

        sheet.cell(row=1, column=1, value='事件ID')
        sheet.cell(row=1, column=2, value='第1周(20180205~20180211)')
        sheet.cell(row=1, column=3, value='第2周(20180212~20180218)')
        sheet.cell(row=1, column=4, value='第3周(20180219~20180225)')
        sheet.cell(row=1, column=5, value='第4周(20180226~20180304)')
        sheet.cell(row=1, column=6, value='第5周(20180305~20180311)')
        sheet.cell(row=1, column=7, value='第6周(20180312~20180318)')
        sheet.cell(row=1, column=8, value='第6周(20180319~20180325)')
        sheet.cell(row=1, column=9, value='第6周(20180326~20180401)')
        sheet.cell(row=1, column=10, value='第6周(20180402~20180404)')

        for row in range(2, len(event_list) + 2):
            sheet.cell(row=row, column=1, value=event_list[row - 2])

        for index in range(1, 10):
            for event_index in range(0, len(event_list)):
                if event_list[event_index] in total_map[index]:
                    sheet.cell(row=event_index + 2, column=index + 1, value=total_map[index][event_list[event_index]])
                else:
                    sheet.cell(row=event_index + 2, column=index + 1, value=0)

        wb.save(title + self.file_date_str + '.xlsx')

    def get_data(self, app_id, start_date, end_date, os):
        # 预获取总数据量
        response_init = requests.get(self.base_url + self.url.format(appid=app_id, sd=start_date, ed=end_date, limit=self.limit, os=os), cookies=self.cookies)
        json_data_init = json.loads(response_init.text)
        limit = int(json_data_init['total'])

        # 获取真实数据
        url = self.base_url + self.url.format(appid=app_id, sd=start_date, ed=end_date, limit=limit, os=os)
        print(start_date + '~' + end_date + '-url:', url)
        response = requests.get(url, cookies=self.cookies)

        return json.loads(response.text)['stats']  # 返回解析的数据


if __name__ == '__main__':
    tjMd = TjMd()
    tjMd.handle()
