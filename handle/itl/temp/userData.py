"""
用户累积数据
"""
import datetime
import json

import openpyxl
import requests


class UserData:
    def __init__(self):
        self.today = datetime.datetime.now()
        self.today_int = int(self.today.strftime('%Y%m%d'))
        self.file_date_str = '_' + str(self.today_int)

        self.base_url = 'http://admin.itianluo.cn/index.php?s=/StoreLogintotallog/logintotallogsRead&_dc=1521628844635'
        self.cookies = {
            'PHPSESSID': 'ibrd42cqe4ci2i69p3s745jk55',
            'UM_distinctid': '161f54da4845fc-0eca83f19c2a67-32667b04-fa000-161f54da48566'
        }
        self.param = {
            'zone_id': 2,
            'date_begin': '2015-01-01',
            'date_end': '1970-01-01',
            'page': 1,
            'start': 1,
            'limit': 1,
            'sort': '[{"property":"id","direction":"desc"}]',
        }

        self.zone = [
            {'id': 2, 'name': '景江苑'},
            {'id': 76, 'name': '天水家园'},
            {'id': 82, 'name': '浙地人家'},
            {'id': 84, 'name': '南城景园'}
        ]

    def h(self):
        wb = openpyxl.Workbook()
        for zone in self.zone:
            self.param['zone_id'] = zone['id']

            response = requests.post(self.base_url, data=self.param, cookies=self.cookies)
            result_data = json.loads(response.text)

            data = result_data['logintotallogs']

            sheet = wb.create_sheet(zone['name'], 0)
            # sheet = wb.active
            sheet.title = zone['name']

            sheet.cell(row=1, column=1, value='日期')
            sheet.cell(row=1, column=2, value='IOS登录用户数')
            sheet.cell(row=1, column=3, value='安卓登录用户数')

            row_index = 2
            for index in range(0, len(data)):
                ios = int(data[index + 1]['iosUserTotal']) if index + 1 < len(data) else 0
                andr = int(data[index + 1]['andUserTotal']) if index + 1 < len(data) else 0

                sheet.cell(row=row_index, column=1, value=data[index]['create_at'])
                sheet.cell(row=row_index, column=2, value=int(data[index]['iosUserTotal']) - ios)
                sheet.cell(row=row_index, column=3, value=int(data[index]['andUserTotal']) - andr)
                row_index += 1

        wb.save('累计登陆用户数' + self.file_date_str + '.xlsx')


if __name__ == '__main__':
    ud = UserData()
    ud.h()
